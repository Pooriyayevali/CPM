import openpyxl


def calculate_cpm(file_path):
    # باز کردن فایل اکسل
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # مشخصات فعالیت‌ها و وابستگی‌ها
    activities = {}
    dependencies = {}
    head = 1
    # خواندن اطلاعات از سلول‌ها و پر کردن دیکشنری‌ها
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        activity, depend, duration = row
        if not activity:
            head += 1
            continue
        dependencies[activity] = [] if not depend else [int(dep.strip()) for dep in str(depend).split(',')]
        activities[activity] = {'duration': duration, 'early_start': 0, 'early_finish': 0,
                                'late_start': None, 'late_finish': None, 'free_respite':None}


    # محاسبه EPO (Early Start, Early Finish)
    for activity in activities:
        if not dependencies[activity]:
            activities[activity]['early_start'] = 0
            activities[activity]['early_finish'] = activities[activity]['duration']
        else:
            activities[activity]['early_start'] = max(
                [activities[depend]['early_finish'] for depend in dependencies[activity]])
            activities[activity]['early_finish'] = activities[activity]['early_start'] + activities[activity]['duration']
    # def print_cpm():
    #     for x in activities:
    #         print(x)
    #         for y in activities[x]:
    #             print(y, ':', activities[x][y])

    # محاسبه LPO (Late Start, Late Finish)
    # مشخص کردن آخرین فعالیت
    end_activity = max(activities, key=lambda x: activities[x]['early_finish'])
    activities[end_activity]['late_finish'] = activities[end_activity]['early_finish']
    activities[end_activity]['late_start'] = activities[end_activity]['late_finish'] - activities[end_activity]['duration']

    # مشخص کردن باقی فعالیت ها بصورت معکوس
    for activity in reversed(list(activities.keys())):
        if activity != end_activity:
            post_depend = []
            for act, deps in reversed(dependencies.items()):
                for dep in deps:
                    if dep == activity:
                        post_depend.append(act)
            activities[activity]['late_finish'] = activities[end_activity]['late_finish'] if not post_depend else\
                min([activities[depend]['late_start'] for depend in post_depend])



            # محاسبه فرجه آزاد با بررسی فعالیت های بعدی وابسته به فعالیت در حال بررسی
            next_dep = min([activities[depend]['early_start'] for depend in post_depend])  if post_depend else activities[end_activity]['late_finish']
            activities[activity]['free_respite'] = next_dep - activities[activity]['early_finish']
            activities[activity]['late_start'] = activities[activity]['late_finish'] - activities[activity]['duration']
        else:
            # فرجه آزاد فعالیت آخر صفر است
            activities[activity]['free_respite'] = 0


    # محاسبه مسیر بحرانی
    critical_path = [activity for activity in activities if activities[activity]['early_start'] ==
                     activities[activity]['late_start'] and activities[activity]['early_finish'] ==
                     activities[activity]['late_finish']]

    # ردیف اول نام ستون ها در اکسل
    sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
    sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
    sheet.cell(row=1, column=4, value="زود ترین")
    sheet.cell(row=1, column=6, value="دیر ترین")
    sheet.cell(row=1, column=8, value="فرجه")

    # ردیف دوم نام ستون ها در اکسل
    sheet.cell(row=2, column=4, value="EPO")
    sheet.cell(row=2, column=5, value="EFT")
    sheet.cell(row=2, column=6, value="LPO")
    sheet.cell(row=2, column=7, value="LFT")
    sheet.cell(row=2, column=8, value="کلی")
    sheet.cell(row=2, column=9, value="آزاد")


    # پرکردن جدول با اطلاعات دیکشنری ها
    for idx, activity in enumerate(activities, start=3):
        sheet.cell(row=idx, column=4, value=activities[activity]['early_start'])
        sheet.cell(row=idx, column=5, value=activities[activity]['early_finish'])
        sheet.cell(row=idx, column=6, value=activities[activity]['late_start'])
        sheet.cell(row=idx, column=7, value=activities[activity]['late_finish'])
        sheet.cell(row=idx, column=8, value=activities[activity]['late_start'] - activities[activity]['early_start'])
        sheet.cell(row=idx, column=9, value=activities[activity]['free_respite'])

    # ذخیره فایل اکسل
    workbook.save(file_path)

# مثال: تابع را با فایل اکسل مورد نظر فراخوانی کنید
calculate_cpm("project.xlsx")